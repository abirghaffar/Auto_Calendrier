# -*- coding: utf-8 -*-
"""
Created on Sun Mar 24 20:47:41 2024

@author: abir
"""  
#Importation des bibliothéques 
import os
import openpyxl
from datetime import datetime, timedelta
import re
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pandas as pd
#Déclaration des variables 
values_Table= range(8, 34)
plage_colonnes=range(6,16)
# Définition des fonctions utilisées 
def liste_dates(start_date, num_days):
    # Convertir la date de départ en un objet datetime
    current_date = datetime.strptime(start_date, '%Y-%m-%d')
    dates = []
    # Boucler pour générer les dates pour le nombre de jours spécifié
    for _ in range(num_days):
        # Exclure les week-ends (samedi et dimanche)
        if current_date.weekday() < 5:  # 0: lundi, 1: mardi, ..., 5: samedi, 6: dimanche
            dates.append(current_date.strftime('%Y-%m-%d'))
            dates.append(current_date.strftime('%Y-%m-%d'))
        # Passer au jour suivant
        current_date += timedelta(days=1)
    return dates

# Load the existing Excel file
existing_file = "C:/Users/abir/Documents/Projet _ Conception numerique/EDT_SIGMA.xlsx"
wb_existing = openpyxl.load_workbook(existing_file, data_only=True)

# Select the specific sheet by name
sheet_existing = wb_existing["M1 2324"]
# "M1 2324" c'est le nom de feuille Excel contenant le calendrier des cours de Master SIGMA 1 

# Create a new Excel file
new_file_name = 'output.xlsx'
new_file_path = os.path.join(os.path.dirname(existing_file), new_file_name)
wb_new = openpyxl.Workbook()

# Select the active sheet of the new file
sheet_new = wb_new.active

# Add headers to the new file
headers = ["Start_Date","End_Date","Start_Time", "End_Time","Subject", "Comment","Location"]
for col_num, header in enumerate(headers, start=1):
    sheet_new.cell(row=1, column=col_num, value=header)

# Iterate through the specified range F83
pattern1 = r'\b\d{1,2}h\d{2}\b-\d{1,2}h\d{2}\b'
pattern2 = r'\b\d{1,2}h-\d{1,2}h\b'
time_ranges = ['08h30 - 12h30', '13h30 - 17h30']

for idx, row in enumerate(values_Table):
    for col in plage_colonnes :
        # Read the value and comment of each cell
        Unite = sheet_existing.cell(row=row, column=col).value
        comment = sheet_existing.cell(row=row, column=col).comment.text if sheet_existing.cell(row=row, column=col).comment else None
        # Write the date, value, and comment to the new Excel file
        if comment:
            plages_horaires = re.findall(pattern1, comment)
            plages_horaires_str = ', '.join(plages_horaires)
            if not(plages_horaires):
                plages_horaires = re.findall(pattern2, comment)
                plages_horaires_str = ', '.join(plages_horaires)
                if not(plages_horaires):
                    time_range_index = 0 if col % 2 == 0 else 1
                    plages_horaires_str = time_ranges[time_range_index]   
            if plages_horaires:     
                Start_Time = plages_horaires[0].split("-")[0].split(" ")[0]
                End_Time=plages_horaires[0].split("-")[-1].split(" ")[-1]
        else:
            time_range_index = 0 if col % 2 == 0 else 1
            plages_horaires = time_ranges[time_range_index]
            Start_Time = plages_horaires.split("-")[0].split(" ")[0]
            End_Time=plages_horaires.split("-")[-1].split(" ")[-1]
        # Ajouter cette chaîne à la liste d'éléments à écrire dans la feuille Excel
            sheet_new.append(["","", Start_Time, End_Time, Unite, comment, " "])

# Exemple d'utilisation :
start_date = '2023-09-18'  # Date de départ
num_days = 365 # Nombre de jours pour générer
Start_Date= liste_dates(start_date, num_days)
End_Date=liste_dates(start_date, num_days)

for idx, row in enumerate(sheet_new, start=2):
    # Insert the value at the beginning of the row
    sheet_new.cell(row=idx, column=1, value=Start_Date[idx - 2])
    sheet_new.cell(row=idx, column=2, value=End_Date[idx - 2])
# Save the new Excel file in the same folder
wb_new.save("output.xls")
#Etape 2 : Passage de format xls  vers ical 
# Chemin vers votre fichier Excel
excel_path = "C:/Users/abir/Documents/Projet _ Conception numerique/output.xls"

# Charger le fichier Excel
df = pd.read_excel(excel_path)
df.index
df.Start_Date = pd.Series(df.Start_Date).fillna(method='ffill')
df.End_Date= pd.Series(df.End_Date).fillna(method='ffill')
# Chemin vers le fichier credentials.json
creds = None
creds_path ="C:/Users/abir/Downloads/credantial.json.json"
# Si nécessaire, changez le scope selon vos besoins
SCOPES = ['https://www.googleapis.com/auth/calendar']

# Authentification et création du service
flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
creds = flow.run_local_server(port=0)
service = build('calendar', 'v3', credentials=creds)

# Création des événements à partir du DataFrame
for i in range(len(df)):
    row = df.iloc[i]
    print(row['Start_Date'])
    print(datetime.strptime(str(row['Start_Date'], '%Y-%m-%d %H:%M:%S')))
    print(row['End_Date'])
    print(datetime.strptime(str(row['End_Date'], '%Y-%m-%d %H:%M:%S')))
    print(row['Start_Time'])
    print(datetime.strptime(str(row['Start_Time'], '%H:%M:%S')))
    print(row['End_Time'])
    print(datetime.strptime(str(row['End_Time'], '%H:%M:%S')))
   
    # Conversion des chaînes de dates en objets datetime
    start_datetime = datetime.strptime(str(row['Start_Date'].strftime('%Y-%m-%d') )
                                       + ' ' 
                                       + (str(row['Start_Time'].strftime('%H:%M:%S'))),
                                       '%Y-%m-%d %H:%M:%S')
    
    end_datetime = datetime.strptime(str(row['End_Date'].strftime('%Y-%m-%d') )
                                        + ' ' 
                                        + (str(row['End_Time'].strftime('%H:%M:%S'))),
                                        '%Y-%m-%d %H:%M:%S')
                                                                
    start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %Hh%M')
    end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %Hh%M')         
event = {
      'summary': row['Subject'],
      'location': row['Location'],
      'description': row['Comment'],
      'start': {
        'dateTime': start_datetime.isoformat(),
        'timeZone': 'Europe/Paris',
      },
      'end': {
        'dateTime': end_datetime.isoformat(),
        'timeZone': 'Europe/Paris',
      },
    }
#Pour insérer les évenements dans votre calendrier , il faut utiliser l'identifiant associé à cet Agenda 
event = service.events().insert(calendarId='87d724daf778aca4455c51c24cc53bc0e94e99beae740a9cf4524b0dcf3de8f7@group.calendar.google.com', body=event).execute()
#Affichage des évenements dans l'Agenda Google dédié
print('Événement créé : %s' % (event.get('htmlLink')))
    