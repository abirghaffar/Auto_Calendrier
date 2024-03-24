# -*- coding: utf-8 -*-
"""
Created on Fri Mar 22 10:24:18 2024

@author: abir
"""
#Importation des bibliothéques 
import os
import openpyxl
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pandas as pd
import os.path
import google.auth
#parameters
table_range = range(8, 34)

# Load the existing Excel file
existing_file = "C:/Users/abir/Documents/Projet _ Conception numerique/EDT_SIGMA.xlsx"
wb_existing = openpyxl.load_workbook(existing_file, data_only=True)

# Select the specific sheet by name
sheet_existing = wb_existing["M1 2324"]
# Create a new Excel file
new_file_name = 'output.xlsx'
new_file_path = os.path.join(os.path.dirname(existing_file), new_file_name)
wb_new = openpyxl.Workbook()

# Select the active sheet of the new file
sheet_new = wb_new.active

# Add headers to the new file
headers = ["Date","Times", "Subject", "Comment"]
for col_num, header in enumerate(headers, start=1):
    sheet_new.cell(row=1, column=col_num, value=header)

def add_date_column(start_date, filename, sheetname, header=True):

    # Get the maximum number of rows in the sheet
    max_row = sheet_existing.max_row

    # If header is True, add the header row
    if header:
        sheet_existing["A1"] = "Date"
        max_row += 1

    # Generate a list of dates starting from the given date
    date_list = [start_date + timedelta(days=x) for x in range(max_row-1)]

    # Add the date column
    for row, date in zip(range(2, max_row+1), date_list):
        sheet_new.cell(row=row, column=1, value=date.strftime("%Y-%m-%d %H:%M:%S"))

# Initialize the start date
start_date = datetime(2023, 9, 18, 0, 0, 0)
# Call the function
add_date_column(start_date,"EDT_SIGMA.xlsx", "M1 2324", header=True)
# Iterate through the specified range F83
pattern1 = r'\b\d{1,2}h\d{2}\b-\d{1,2}h\d{2}\b'
pattern2 = r'\b\d{1,2}h-\d{1,2}h\b'
time_ranges = ['08h30 - 12h30', '13h30 - 17h30']

for idx, row in enumerate(range(8, 34)):
    for col in range(6, 16):
        # Read the value and comment of each cell
        Unite = sheet_existing.cell(row=row, column=col).value
        comment = sheet_existing.cell(row=row, column=col).comment.text if sheet_existing.cell(row=row, column=col).comment else None
        # Write the date, value, and comment to the new Excel file
        if comment:
            plages_horaires = re.findall(pattern1, comment)
            plages_horaires_str = ', '.join(plages_horaires)
            if not(plages_horaires):
                plages_horaires = re.findall(pattern2, comment)
                plages_horaires_str = ', '.join(plages_horaires)
                if not(plages_horaires):
                    time_range_index = 0 if col % 2 == 0 else 1
                    plages_horaires_str = time_ranges[time_range_index]        
        else:
            time_range_index = 0 if col % 2 == 0 else 1
            plages_horaires_str = time_ranges[time_range_index]

        # Addition of the liste of the date, the unit and the comment to the new Excel sheet 
        sheet_new.append([plages_horaires_str, Unite, comment])
for row in range(2,249):
    duree= sheet_new.cell(row=row,column=2)
    print(duree)
    start_time, end_time=duree.split("-")
    print(start_time, end_time)
#Convetion of start_time and end_time to datetime objects 
start_time_objet = datetime.datetime.strptime(start_time,"%H:%M")
end_time_objet = datetime.datetime.strptime(start_time,"%H:%M")
print(start_time_objet)
print(end_time_objet)
# Save the new Excel file in the same folder
wb_new.save('C:/Users/abir/Documents/Projet _ Conception numerique/output.xlsx')

# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 21:45:04 2024

@author: abir
"""
# path to the Excel file already created 
excel_path = "C:/Users/abir/Documents/Projet _ Conception numerique/output.xlsx"

# loading the Excel file 
df = pd.read_excel(excel_path)
df.index
df.Date = pd.Series(df.Date).fillna(method='ffill')
# path to the credentials.json file 
creds = None
creds_path = "C:/Users/abir/Documents/Projet _ Conception numerique/credentials.json.json"

# if you need, change the scope as you want depending to your needs 
SCOPES = ['https://www.googleapis.com/auth/calendar']

# Authentication and creation of service 
flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
creds = flow.run_local_server(port=0)
service = build('calendar', 'v3', credentials=creds)

# Events creation from DataFrame
for i in range(len(df)):
    row = df.iloc[i]
    print(row['Date'])
    print(datetime.strptime(str(row['Date'], '%Y-%m-%d %H:%M:%S')))
    print(row['Times'])
    print(datetime.strptime(str(row['Times'], '%H:%M:%S')))
   
    # Conversion from dates as string to datetime object 
    start_datetime = datetime.strptime(str(row['Date'].strftime('%Y-%m-%d') )
                                       + ' ' 
                                       + (str(row['Times'].strftime('%H:%M:%S'))),
                                       '%Y-%m-%d %H:%M:%S')
                                   
    
event = {
      'summary': row['Unit'],
   
      'description': row['Commentaire'],
      'start': {
        'dateTime': start_datetime.isoformat(),
        'timeZone': 'Europe/Paris',}
     
    }

event = service.events().insert(calendarId='1dd078c465b410198c221ccf5fcf01110cb814125d5d08230c8cf107c2981ac2@group.calendar.google.com', body=event).execute()

print('Événement créé : %s' % (event.get('htmlLink')))

    # -*- coding: utf-8 -*-
"""
Created on Sun Mar 24 14:02:53 2024

@author: abir
"""

os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "C:/Users/abir/Documents/Projet _ Conception numerique/Credentials.json (2).json"

# Load credentials and create an API client
creds, project = google.auth.default(scopes=["https://www.googleapis.com/auth/calendar.readonly"])
service = build("calendar", "v3", credentials=creds)

# Set up the Excel file
wb = openpyxl.Workbook()
ws = wb.active

# Get the calendar events
events_result = service.events().list(calendarId="primary").execute()
events = events_result.get("items", [])

# Write the events to the Excel file
for i, event in enumerate(events):
    start = event["start"].get("dateTime", event["start"].get("date"))
    summary = event["summary"]
    ws.cell(row=i+1, column=1, value=start)
    ws.cell(row=i+1, column=2, value=summary)

# Save the Excel file
wb.save("calendar_M1.xlsx")
